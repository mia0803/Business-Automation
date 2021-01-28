# "Partial codes" of business automation program 

# Paste Data Frame to existing file
import pandas as pd
from openpyxl import load_workbook

export = ""
book = load_workbook(export)
writer = pd.ExcelWriter(export, engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}

frame.to_excel(writer, sheet_name='Main', startcol= 25, startrow=10, index = False, header= False)

writer.save()


# Invoice Date & Service Period
from openpyxl import load_workbook

wb = load_workbook(export)
sheet = wb['Main']
wb2 = load_workbook(backup)
sheet2 = wb2['Invoice Summary']

start=11
for each in customer:
    sheet['L'+str(start)] = date
    sheet['Q'+str(start)] = period
    start+=1

sheet['AC2'] = sheet2['E6'].value
sheet['AC3'] = sheet2['E7'].value
    
wb.save(export)


# Save each sheet as pdf file
import win32com.client

o = win32com.client.Dispatch("Excel.Application")
o.Visible = False

wb = o.Workbooks.Open(file)

for sheet in range(length):
    if sheet > 1:      # Skip first two "invoice", "summary" sheets
        ws_index = sheet+1 
        path_to_pdf = path+"{}.pdf".format(namelist[sheet].replace(" ","_"))

        wb.WorkSheets(ws_index).Select()

        wb.ActiveSheet.ExportAsFixedFormat(0, path_to_pdf)


# Merge PDF Files
import PyPDF2
import os

invoice=[]
for filename in os.listdir(invoicepath):
    invoice.append(filename)
    
backup=[]
for filename in os.listdir(backuppath):
    backup.append(filename)

h=0
for file in invoice:
    pdfFileObj = open(backuppath + backup[h], 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    pageObj = pdfReader.getPage(0)
    text=pageObj.extractText()
    textlist=text.split('\n')
    code=textlist[5].strip("        ")
    if code == file[-6:-4]:
        merger = PyPDF2.PdfFileMerger()
        merger.append(invoicepath + file)
        merger.append(backuppath + backup[h])
        merger.write(newpath + file)
        merger.close()
    else:
        print('Warning: Invoice file "', end='')
        print(file, end='')
        print('" does not match with Backup file code "', end='')
        print(code, end='')
        print('". Please check.')
    h+=1

pdfFileObj.close()


# Sending email
for each in invoice:
    code=each[7:9]
    if not code in email:
        print("There is no email list for invoice " + str(code) )
    elif len(email[code])==1:
        to = email[code][0] # String
        receiver_email = to # It is okay to have list
        subject =  each[:9] + " Wheelchair Service Invoice (TBIT/LAX)"
        message = MIMEMultipart()
        # Below should be string
        message["From"] = sender_email
        message["To"] = to
        message["Subject"] = subject
        
        message.attach(MIMEText(html, "html"))
        
        # File attachment
        with open(invoicepath + each, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())

        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename = each) 
        message.attach(part)

        text = message.as_string()
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email, text)
    
    else:
        to = email[code][0] # String
        cc = email[code][1:] # List

        receiver_email =  [to] + cc # It is okay to have list
        subject =  each[:9] + " Wheelchair Service Invoice (TBIT/LAX)"
        message = MIMEMultipart()
        # Below should be string
        message["From"] = sender_email
        message["To"] = to
        message["Subject"] = subject
        message["Cc"] = "%s" % ",".join(cc) 
        
        message.attach(MIMEText(html, "html"))
        
        # File attachment
        with open(invoicepath + each, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())

        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment', filename = each) 
        message.attach(part)

        text = message.as_string()
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_email, text)



